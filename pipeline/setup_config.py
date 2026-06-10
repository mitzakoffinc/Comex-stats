"""
One-off: ensure Config/config.xlsx has the 'target_urfs' sheet required by
notebooks/07_ncm_scoring.py (route-scoring criterion).

Codes are 7-digit zero-padded URF codes from Data/References/URF.csv.
Safe to re-run — replaces the sheet if it already exists.

Run from project root:
    python pipeline/setup_config.py
"""

import sys
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))
from config import CONFIG_XLSX

TARGET_URFS = [
    ("0817600", "Aeroporto Internacional de Sao Paulo/Guarulhos (GRU)"),
    ("0817700", "Aeroporto Internacional de Viracopos (VCP)"),
    ("0817800", "Porto de Santos"),
    ("0810600", "Santos (legacy URF code)"),
    ("0811100", "Guarulhos (legacy URF code)"),
]


def main() -> None:
    wb = load_workbook(CONFIG_XLSX)

    if "target_urfs" in wb.sheetnames:
        del wb["target_urfs"]
        print("Existing 'target_urfs' sheet replaced.")

    ws = wb.create_sheet("target_urfs")
    ws.append(["co_urf", "descricao"])
    for code, desc in TARGET_URFS:
        # Write as text so Excel never strips the leading zero
        cell = ws.cell(row=ws.max_row + 1, column=1, value=code)
        cell.number_format = "@"
        ws.cell(row=ws.max_row, column=2, value=desc)

    wb.save(CONFIG_XLSX)
    print(f"Sheet 'target_urfs' written to {CONFIG_XLSX}")
    print(f"Sheets now: {wb.sheetnames}")
    for code, desc in TARGET_URFS:
        print(f"  {code} — {desc}")


if __name__ == "__main__":
    main()
